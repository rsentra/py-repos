# -*- coding: utf-8 -*-
"""
Created on Sun Apr  4 16:34:59 2021
https://openpyxl.readthedocs.io/en/stable/tutorial.html
@author: 윤제열
0. file 
  - wb = load_workbook("sample_formula.xlsx", data_only=True)
   파일을 읽으면 cell.value에는 수식이 들어있음, 값으로 읽으려면 
    data_only=True 로 읽어야
1. cell참조방법
  - ws['a1']
  - ws.cell(row,col)
  - ws['a']  ws['a:c'] : 컬럼전체
  - ws[2]  ws[1:10] : row전체
  - ws['a1:d10'] : 범위
  - row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
2.cell속성
  - cell.value
  - cell.font
  - cell.number_format
  - cell.border
3.범위 반복
  - ws.iter_rows(min_col=?, max_col=?, min_row=?, max_row=?)
  - ws.iter_cols(min_col=?, max_col=?, min_row=?, max_row=?)
   iter_rows, iter_cols는  values_only=True 옵션으로 value만 읽을수 있음
  - ws.rows 
  - ws.columns
4. 삽입,삭제
  - ws.insert_rows(3, 2) # 3번째 행앞에 2행 추가
  - ws.insert_cols(2, 3) # B번째 열로부터 3열 추가
  - ws.delete_rows(3, 2)
  - ws.delete_cols(2, 3)
  - ws.move_range("C1:C11", rows=5, cols=-1)
5. cell 참조방법들
  - ws.cell(row,col).coordinate 는 'A1' 형식의 주소반환
  - str(chr(65)) => 'A', 'B'는  chr(66)으로 하면 됨
  - str_rng = f'A{r}:D{r}'  => r=10 이라면 'A10:D10'
    sum(str_rng)
6. cell style
  - ws.column_dimensions["A"].width = 5
  - ws.row_dimensions[1].height = 50
  - ws['a1'].font = Font(color="FF0000", italic=True, bold=True) # 글자 색은 빨갛게, 이탤릭, 두껍게 적용
  - ws['b1'].font = Font(color="CC33FF", name="Arial", strike=True) # 폰트를 Arial 로 설정, 취소선 적용
  - ws['c1'].font = Font(color="0000FF", size=20, underline="single") # 글자 크기를 20으로, 밑줄 적용
  # 90 점 넘는 셀에 대해서 초록색으로 적용
  for row in ws.rows:
      for cell in row:
        # 각 cell 에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # center, left, right, top, bottom

        if cell.column == 1: # A 번호열은 제외
            continue

        # cell 이 정수형 데이터이고 90 점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로 설정
            cell.font = Font(color="FF0000") # 폰트 색상 변경

   # 틀 고정
     ws.freeze_panes = "B2" # B2 기준으로 틀 고정

  7. formula
   - ws["A6"] = "=SUM(A4:A5)"
   
  8. merge cells 
   - ws.merge_cells("B2:D2") # B2 부터 D2 까지 합치겠음
   - ws["B2"].value = "Merged Cell"
   - ws.unmerge_cells("B2:D2")
   
   ** cell속성은 개별cell별로 설정해야 함 <-반복문 이용
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors

ym_nm = '10월'
wb = load_workbook("2020-10 롯데정산결과(KJ).xlsx")
ws = wb.active

wb1 = load_workbook("2020-09 롯데정산결과(KJ).xlsx",data_only=True)
ws1 = wb1.active

def find_row(ws,val):
    # for row in ws['a'].iter_rows():
    #     for cell in row:
    for cell in ws['a']:
        if val == cell.value:
            return cell.row

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

big_blue_text = Font(color=colors.BLUE, size=14)
double_border_side = Side(border_style="double")
square_border = Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)

# 위치 찾기
r = find_row(ws,"합계")
c = [ col.column for col in ws[1] if col.value =='이자'][0]
이자 = ws.cell(r,c).value
ws.cell(r,c).font = big_blue_text

# -- 전월 파일에서 
r = find_row(ws1, "이자 정산표")
# 요약표
sum_tbl = ws1[r:ws1.max_row]

# 행 위치를 같게 만듬
r_cnt = r - ws.max_row - 1
ws.insert_rows(ws.max_row + 1,r_cnt)
ws.cell(ws.max_row+1,1).value = " "
# for row in ws[ws.max_row+1: ws.max_row + r_cnt]:
#     row[0].value = " "


# 요약표를 복사
# for row in sum_tbl:
for row in ws1.iter_rows(min_row = r,values_only = True):    
    ws.append([c for c in row])


# 직전 최종월 column
r = find_row(ws, "이자 정산표")
c = [ cell.column for cell in ws[r+3] if cell.value is not None][-1]
# 직전 최종월 row
for cell in ws.iter_rows(min_row=r+3, max_row = ws.max_row-1, min_col=c, max_col=c):
    if cell[0].value:
        mx_r = cell[0].row

if mx_r == ws.max_row-1:
    ws.insert_rows(ws.max_row)
    
for cell in ws.iter_rows(min_row = r+3, max_row = ws.max_row-1, min_col=c, max_col = c):
    rr = cell[0].row
    ws.cell(rr, c+1).value =  ws.cell(rr, c).value

# 당월 제목
ws.cell(r+2,c+1).value = ym_nm 
ws.cell(mx_r + 1,1).value = ym_nm 
# 당월 이자
int_rng = ws.cell(mx_r + 1,c+1)
int_rng.value = 이자
# 당월 누적합계
# a = f'{str(chr(65+c))}{r+3}'
# b = f'{str(chr(65+c))}{mx_r+1}'
a = ws.cell(r+3,c+1).coordinate
b = ws.cell(mx_r+1,c+1).coordinate
 
ws.cell(ws.max_row, c+1).value  = f'=sum({a}:{b})'

# style 
int_rng.font = big_blue_text
ws.cell(ws.max_row, c+1).font = big_blue_text

# coordinate는 'a1' 스타일 주소 
a = ws.cell(r+2,1).coordinate
b = ws.cell(ws.max_row,c+1).coordinate
str_rng = f"{a}:{b}"
set_border(ws, str_rng)

str_rng = f"A:{str(chr(65+c+1))}"
for row in ws[str_rng]:
    for cell in row:
         cell.number_format = '#,##0'


wb.save("tets_kj.xlsx")
print('completed')